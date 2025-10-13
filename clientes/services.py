# clientes/services.py
from __future__ import annotations
from typing import Optional, Iterable
import re
from datetime import date
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.db import transaction
import secrets
from datetime import timedelta
from django.utils import timezone
from django.core.exceptions import ValidationError

from .models import Cliente
# clientes/services.py
from datetime import timedelta
import secrets
from django.conf import settings
from django.utils import timezone
from django.urls import reverse

from notificacoes.emails import send_email_html
# clientes/services.py


_ONLY_DIGITS = re.compile(r"\D+")

def clean_doc(s: str) -> str:
    return _ONLY_DIGITS.sub("", (s or ""))

def paginar(qs, page: int = 1, per_page: int = 20):
    p = Paginator(qs, per_page)
    try:
        return p.page(page)
    except PageNotAnInteger:
        return p.page(1)
    except EmptyPage:
        return p.page(p.num_pages if page > 1 else 1)

def buscar_clientes(q: str = "", ativos: Optional[bool] = None, condominio: Optional[int] = None):
    qs = Cliente.objects.all()
    if condominio:
        qs = qs.filter(condominio_id=condominio)
    if q:
        q_clean = clean_doc(q)
        cond = Q(nome_razao__icontains=q) | Q(email__icontains=q)
        if q_clean:
            cond |= Q(cpf_cnpj__icontains=q_clean)
        qs = qs.filter(cond)
    if ativos is not None:
        qs = qs.filter(ativo=ativos)
    return qs.order_by("nome_razao", "id")

# clientes/services.py
from django.core.exceptions import ValidationError

ACEITE_TTL_DIAS = 7  # expira em 7 dias, ajuste se quiser

def _gerar_token_unico() -> str:
    # 32 bytes ~ 43 chars urlsafe, suficiente. Se quiser menor, reduza.
    return secrets.token_urlsafe(32)

def criar_cliente(data: dict) -> Cliente:
    data = {**data}
    data["cpf_cnpj"] = clean_doc(data.get("cpf_cnpj", ""))
    if not data["cpf_cnpj"]:
        raise ValidationError("CPF/CNPJ é obrigatório.")
    if not data.get("condominio"):
        raise ValidationError("Condomínio é obrigatório para o cadastro do cliente.")

    # 1) Sempre inicia inativo
    data["ativo"] = False

    # 2) Gera token e prazo
    token = _gerar_token_unico()
    expires = timezone.now() + timedelta(days=ACEITE_TTL_DIAS)
    data["aceite_token"] = token
    data["aceite_expires_at"] = expires

    # 3) Cria o cliente
    c = Cliente.objects.create(**data)

    # 4) Envia e-mail de aceite (template + link)
    enviar_email_aceite(c)  # já existe no projeto — só garantir que usa o token/URL corretos

    return c


@transaction.atomic
def atualizar_cliente(cliente_id: int, data: dict) -> Cliente:
    print('atualizando cliente')
    c = Cliente.objects.filter(id=cliente_id).first()
    if not c:
        raise ObjectDoesNotExist("Cliente não encontrado.")
    if "cpf_cnpj" in data:
        data["cpf_cnpj"] = clean_doc(data["cpf_cnpj"])
        if not data["cpf_cnpj"]:
            raise ValidationError("CPF/CNPJ inválido.")
    for k, v in data.items():
        setattr(c, k, v)
    c.full_clean()
    c.save()
    return c

@transaction.atomic
def ativar(cliente_id: int, ativo: bool = True) -> Cliente:
    c = Cliente.objects.filter(id=cliente_id).first()
    if not c:
        raise ObjectDoesNotExist("Cliente não encontrado.")
    c.ativo = bool(ativo)
    c.save(update_fields=["ativo", "updated_at"])
    return c

# ==== Importação/Exportação Excel (openpyxl) ====
def importar_excel(file) -> dict:
    """
    Espera um .xlsx com cabeçalhos:
    cpf_cnpj,nome_razao,data_nascimento,telefone_emergencial,telefone_celular,
    cep,numero_id,logradouro,bairro,complemento,municipio,estado,email,ativo
    """
    from openpyxl import load_workbook
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    headers = [str((c.value or "")).strip() for c in next(ws.rows)]
    idx = {h: i for i, h in enumerate(headers)}
    required = ["cpf_cnpj","nome_razao"]
    for r in required:
        if r not in idx:
            raise ValidationError(f"Coluna obrigatória ausente: {r}")

    created = updated = skipped = 0
    for row in ws.iter_rows(min_row=2):
        get = lambda h: (row[idx[h]].value if h in idx else None)
        data = {
            "cpf_cnpj": clean_doc(str(get("cpf_cnpj") or "")),
            "nome_razao": str(get("nome_razao") or "").strip(),
            "data_nascimento": get("data_nascimento"),
            "telefone_emergencial": str(get("telefone_emergencial") or "").strip(),
            "telefone_celular": str(get("telefone_celular") or "").strip(),
            "cep": str(get("cep") or "").strip(),
            "numero_id": str(get("numero_id") or "").strip(),
            "logradouro": str(get("logradouro") or "").strip(),
            "bairro": str(get("bairro") or "").strip(),
            "complemento": str(get("complemento") or "").strip(),
            "municipio": str(get("municipio") or "").strip(),
            "estado": str(get("estado") or "").strip(),
            "email": str(get("email") or "").strip(),
        }
        ativo_val = get("ativo")
        if ativo_val is not None:
            data["ativo"] = bool(ativo_val in (1, "1", True, "True", "true", "SIM", "Sim", "sim"))

        if not data["cpf_cnpj"] or not data["nome_razao"]:
            skipped += 1
            continue

        # upsert por documento
        obj = Cliente.objects.filter(cpf_cnpj=data["cpf_cnpj"]).first()
        if obj:
            for k, v in data.items():
                setattr(obj, k, v)
            obj.save()
            updated += 1
        else:
            Cliente.objects.create(**data)
            created += 1

    return {"created": created, "updated": updated, "skipped": skipped}

def exportar_excel(queryset=None) -> tuple[str, bytes]:
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    qs = queryset or Cliente.objects.all().order_by("nome_razao","id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    headers = ["CPF/CNPJ","Nome","Nascimento","Celular","Emergencial","CEP","Número","Logradouro","Bairro","Complemento","Município","UF","E-mail","Ativo","ID"]
    ws.append(headers)

    for c in qs:
        ws.append([
            c.cpf_cnpj, c.nome_razao,
            c.data_nascimento.isoformat() if c.data_nascimento else "",
            c.telefone_celular, c.telefone_emergencial,
            c.cep, c.numero_id, c.logradouro, c.bairro, c.complemento, c.municipio, c.estado,
            c.email, "SIM" if c.ativo else "NÃO", c.id
        ])

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    bio = BytesIO(); wb.save(bio); bio.seek(0)
    from datetime import datetime
    filename = f"clientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return filename, bio.read()

# services.py
from datetime import timedelta
from django.utils import timezone
from django.conf import settings
from django.urls import reverse
import secrets

ACEITE_TTL_DIAS = 7

def gerar_link_aceite(cliente) -> str:
    token = secrets.token_urlsafe(32)
    cliente.aceite_token = token
    cliente.aceite_expires_at = timezone.now() + timedelta(days=ACEITE_TTL_DIAS)
    cliente.save(update_fields=["aceite_token", "aceite_expires_at"])
    base = getattr(settings, "SITE_URL", "http://127.0.0.1:8000").rstrip("/")
    path = reverse("clientes:aceite", args=[token])
    return f"{base}{path}"

# services.py (trecho)
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings

def _build_abs_url(request, url_path: str) -> str:
    """
    Monta URL absoluta de forma robusta.
    1) Se houver request: usa build_absolute_uri (respeita host/https).
    2) Senão: usa settings.SITE_URL como base.
    """
    if request is not None:
        # Garante path absoluto mesmo que passem uma URL relativa
        return request.build_absolute_uri(url_path)

    base = getattr(settings, "SITE_URL", "").rstrip("/")
    if not base:
        # fallback explícito para não quebrar (melhor configurar SITE_URL!)
        base = "http://127.0.0.1:8000"
    return f"{base}{url_path}"



def enviar_email_aceite(cliente, request=None):
    token = getattr(cliente, "aceite_token", None)
    if not token:
        return

    path_confirm = reverse("clientes:aceite_confirmar", args=[token])
    link_confirm = _build_abs_url(request, path_confirm)  # << usa a função robusta

    ctx = {"cliente": cliente, "link_confirm": link_confirm, "expira_em": cliente.aceite_expires_at}
    html = render_to_string("clientes/email_aceite.html", ctx)
    plain = strip_tags(html)
    send_mail(
        subject="Confirme seu contrato",
        message=plain,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[cliente.email],
        html_message=html,
        fail_silently=False,
    )



def buscar_clientes_api(q: str, limit: int = 15):
    qs = Cliente.objects.filter(ativo=True)
    if q:
        q = q.strip()
        q_clean = "".join(ch for ch in q if ch.isalnum())  # limpa CPF/CNPJ
        qs = qs.filter(
            Q(nome_razao__icontains=q) |
            Q(email__icontains=q) |
            Q(cpf_cnpj__icontains=q_clean)
        )
    return (qs
            .order_by("nome_razao", "id")
            .values("id", "nome_razao", "cpf_cnpj", "email")[:limit])
