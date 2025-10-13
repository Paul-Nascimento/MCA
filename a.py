


# --------- clientes/views.py ---------
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.http import HttpRequest, HttpResponse
from django.views.decorators.http import require_http_methods, require_POST
from django.utils import timezone

from .forms import ClienteForm, ClienteFiltroForm
from .models import Cliente
from . import services as cs


@login_required
def list_clientes(request: HttpRequest):
    f = ClienteFiltroForm(request.GET or None)
    cd = f.cleaned_data if f.is_valid() else {}
    qs = cs.buscar_clientes(
        q=cd.get("q", ""),
        ativos=(None if (cd.get("ativos") in (None, "")) else (cd.get("ativos") == "1")),
        condominio=cd.get("condominio")
    )
    # paginação segura
    page_str = request.GET.get("page", "1")
    try:
        page = max(1, int(page_str))
    except Exception:
        page = 1
    page_obj = cs.paginar(qs, page=page, per_page=20)

    qd = request.GET.copy(); qd.pop("page", None)
    base_qs = qd.urlencode()
    suffix = f"&{base_qs}" if base_qs else ""

    # >>> NOVO: suporte a abrir modal de edição via GET ?edit=<id>
    edit_id = request.GET.get("edit")
    edit_cliente = None
    if edit_id:
        try:
            edit_cliente = Cliente.objects.get(pk=int(edit_id))
        except Exception:
            edit_cliente = None
            messages.error(request, "Cliente para edição não encontrado.")

    return render(request, "clientes/list.html", {
        "filtro_form": f,
        "page_obj": page_obj,
        "suffix": suffix,
        "base_qs": base_qs,
        "cliente_form": ClienteForm(),
        "edit_cliente": edit_cliente,  # <<< usado para autoabrir modal em modo edição
    })


@require_http_methods(["POST"])
@login_required
def create_cliente(request: HttpRequest):
    form = ClienteForm(request.POST)
    if not form.is_valid():
        messages.error(request, f"Dados inválidos: {form.errors.as_json()}")
        return redirect(reverse("clientes:list"))

    try:
        cliente = cs.criar_cliente(form.cleaned_data)
        if not getattr(cliente, "aceite_token", None):
            cs.gerar_link_aceite(cliente)
        cs.enviar_email_aceite(cliente, request=request)
        messages.success(request, "Cliente criado e e-mail de aceite enviado.")
    except Exception as e:
        messages.error(request, f"Erro ao criar: {e}")

    return redirect(reverse("clientes:list"))


@require_http_methods(["POST"])
@login_required
def update_cliente(request: HttpRequest, pk: int):
    form = ClienteForm(request.POST)
    if form.is_valid():
        try:
            cs.atualizar_cliente(pk, form.cleaned_data)
            messages.success(request, "Cliente atualizado.")
        except Exception as e:
            messages.error(request, f"Erro ao atualizar: {e}")
            # Redireciona abrindo modal de edição novamente
            return redirect(f"{reverse('clientes:list')}?edit={pk}")
    else:
        messages.error(request, f"Dados inválidos: {form.errors.as_json()}")
        return redirect(f"{reverse('clientes:list')}?edit={pk}")
    return redirect(reverse("clientes:list"))


@require_http_methods(["POST"])
@login_required
def ativar_cliente(request: HttpRequest, pk: int):
    try:
        ativo = request.POST.get("ativo", "1") == "1"
        cs.ativar(pk, ativo=ativo)
        messages.success(request, "Cliente atualizado.")
    except Exception as e:
        messages.error(request, f"Erro: {e}")
    return redirect(reverse("clientes:list"))


@require_http_methods(["POST"])
@login_required
def importar_excel_view(request: HttpRequest):
    f = request.FILES.get("arquivo")
    if not f:
        messages.error(request, "Selecione um arquivo .xlsx")
        return redirect(reverse("clientes:list"))
    try:
        rel = cs.importar_excel(f)
        messages.success(
            request,
            f"Importação: criados={rel.get('created',0)} "
            f"atualizados={rel.get('updated',0)} "
            f"ignorados={rel.get('skipped',0)}"
        )
    except Exception as e:
        messages.error(request, f"Falha ao importar: {e}")
    return redirect(reverse("clientes:list"))


@login_required
def exportar(request: HttpRequest):
    qs = cs.buscar_clientes(
        q=request.GET.get("q", ""),
        ativos=None if request.GET.get("ativos", "") == "" else request.GET.get("ativos") == "1"
    )
    filename, content = cs.exportar_excel(qs)
    resp = HttpResponse(content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@require_http_methods(["GET", "POST"])  # público
def aceite_contrato(request: HttpRequest, token: str):
    cliente = get_object_or_404(Cliente, aceite_token=token)
    if not getattr(cliente, "aceite_expires_at", None) or timezone.now() > cliente.aceite_expires_at:
        messages.error(request, "Link de confirmação expirado. Solicite um novo convite.")
        return render(request, "clientes/aceite_contrato.html", {"expirado": True, "cliente": cliente})

    if request.method == "POST":
        cliente.ativo = True
        cliente.aceite_confirmado_em = timezone.now()
        cliente.aceite_token = None
        cliente.aceite_expires_at = None
        cliente.save(update_fields=["ativo", "aceite_confirmado_em", "aceite_token", "aceite_expires_at"])
        return render(request, "clientes/aceite_sucesso.html", {"cliente": cliente})

    return render(request, "clientes/aceite_contrato.html", {"cliente": cliente})


@require_POST
@login_required
def toggle_status(request, pk: int):
    c = get_object_or_404(Cliente, pk=pk)
    if not c.ativo and getattr(c, "aceite_token", None):
        messages.info(request, "Cliente aguardando confirmação; altere o status somente após a confirmação ou reenviar o convite.")
        return redirect(reverse("clientes:list"))

    c.ativo = not c.ativo
    c.save(update_fields=["ativo"])
    messages.success(request, ("Cliente ativado." if c.ativo else "Cliente desativado."))
    return redirect(reverse("clientes:list"))



# --------- clientes/services.py ---------
from __future__ import annotations
from typing import Optional
import re
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.db import transaction
import secrets
from datetime import timedelta
from django.utils import timezone

from .models import Cliente
from django.conf import settings
from django.urls import reverse

_ONLY_DIGITS = re.compile(r"\D+")

def clean_doc(s: str) -> str:
    return _ONLY_DIGITS.sub("", (s or ""))


def paginar(qs, page: int = 1, per_page: int = 20):
    p = Paginator(qs, per_page)
    try:
        return p.page(page)
    except PageNotAnInteger:
        return p.page(1)
    except EmptyPage:
        return p.page(p.num_pages if page > 1 else 1)


def buscar_clientes(q: str = "", ativos: Optional[bool] = None, condominio: Optional[int] = None):
    qs = Cliente.objects.all()
    if condominio:
        qs = qs.filter(condominio_id=condominio)
    if q:
        q_clean = clean_doc(q)
        cond = Q(nome_razao__icontains=q) | Q(email__icontains=q)
        if q_clean:
            cond |= Q(cpf_cnpj__icontains=q_clean)
        qs = qs.filter(cond)
    if ativos is not None:
        qs = qs.filter(ativo=ativos)
    return qs.order_by("nome_razao", "id")


ACEITE_TTL_DIAS = 7

def _gerar_token_unico() -> str:
    return secrets.token_urlsafe(32)


def criar_cliente(data: dict) -> Cliente:
    data = {**data}
    data["cpf_cnpj"] = clean_doc(data.get("cpf_cnpj", ""))
    if not data["cpf_cnpj"]:
        raise ValidationError("CPF/CNPJ é obrigatório.")
    if not data.get("condominio"):
        raise ValidationError("Condomínio é obrigatório para o cadastro do cliente.")

    data["ativo"] = False
    token = _gerar_token_unico()
    expires = timezone.now() + timedelta(days=ACEITE_TTL_DIAS)
    data["aceite_token"] = token
    data["aceite_expires_at"] = expires

    c = Cliente.objects.create(**data)
    enviar_email_aceite(c)
    return c


@transaction.atomic
def atualizar_cliente(cliente_id: int, data: dict) -> Cliente:
    c = Cliente.objects.filter(id=cliente_id).first()
    if not c:
        raise ObjectDoesNotExist("Cliente não encontrado.")
    if "cpf_cnpj" in data:
        data["cpf_cnpj"] = clean_doc(data["cpf_cnpj"])
        if not data["cpf_cnpj"]:
            raise ValidationError("CPF/CNPJ inválido.")
    for k, v in data.items():
        setattr(c, k, v)
    c.full_clean()
    c.save()
    return c


@transaction.atomic
def ativar(cliente_id: int, ativo: bool = True) -> Cliente:
    c = Cliente.objects.filter(id=cliente_id).first()
    if not c:
        raise ObjectDoesNotExist("Cliente não encontrado.")
    c.ativo = bool(ativo)
    c.save(update_fields=["ativo", "updated_at"])
    return c


def exportar_excel(queryset=None) -> tuple[str, bytes]:
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    qs = queryset or Cliente.objects.all().order_by("nome_razao","id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    headers = ["CPF/CNPJ","Nome","Nascimento","Celular","Emergencial","CEP","Número","Logradouro","Bairro","Complemento","Município","UF","E-mail","Ativo","ID"]
    ws.append(headers)

    for c in qs:
        ws.append([
            c.cpf_cnpj, c.nome_razao,
            c.data_nascimento.isoformat() if c.data_nascimento else "",
            c.telefone_celular, c.telefone_emergencial,
            c.cep, c.numero_id, c.logradouro, c.bairro, c.complemento, c.municipio, c.estado,
            c.email, "SIM" if c.ativo else "NÃO", c.id
        ])

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    bio = BytesIO(); wb.save(bio); bio.seek(0)
    from datetime import datetime
    filename = f"clientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return filename, bio.read()


# ---- Fluxo de aceite (helpers) ----

def gerar_link_aceite(cliente: Cliente) -> str:
    token = _gerar_token_unico()
    cliente.aceite_token = token
    cliente.aceite_expires_at = timezone.now() + timedelta(days=ACEITE_TTL_DIAS)
    cliente.save(update_fields=["aceite_token","aceite_expires_at"])
    base = getattr(settings, "SITE_URL", "http://127.0.0.1:8000").rstrip("/")
    path = reverse("clientes:aceite", args=[token])
    return f"{base}{path}"


def _build_abs_url(url_path: str) -> str:
    base = getattr(settings, "SITE_URL", "").rstrip("/")
    return f"{base}{url_path}"


def enviar_email_aceite(cliente: Cliente, request=None):
    if not getattr(cliente, 'aceite_token', None):
        return
    path = reverse("clientes:aceite", args=[cliente.aceite_token])
    _ = _build_abs_url(path)
    # Integre com seu envio real de e-mail aqui
    return


# --------- clientes/urls.py ---------
from django.urls import path
from . import views

app_name = "clientes"

urlpatterns = [
    path("clientes/", views.list_clientes, name="list"),
    path("clientes/criar/", views.create_cliente, name="create"),
    path("clientes/<int:pk>/atualizar/", views.update_cliente, name="update"),
    path("clientes/<int:pk>/status/", views.toggle_status, name="toggle_status"),
    path("clientes/aceite/<str:token>/", views.aceite_contrato, name="aceite"),
    path("exportar/", views.exportar, name="exportar"),
    path("<int:pk>/ativar/", views.toggle_status, name="ativar"),
]
