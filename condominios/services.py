from __future__ import annotations
from typing import Optional, Dict, Any, Tuple
from io import BytesIO
from datetime import datetime

from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.db import transaction
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from .models import Condominio, UF_CHOICES

# =====================
# Exceções de domínio
# =====================
class CondominioJaExiste(ValidationError):
    pass

class CondominioNaoEncontrado(ObjectDoesNotExist):
    pass

# =====================
# Helpers
# =====================
def _only_digits(s: Any) -> str:
    return "".join(ch for ch in str(s or "") if ch.isdigit())

def _assert_condominio(obj: Optional[Condominio], msg="Condomínio não encontrado"):
    if obj is None:
        raise CondominioNaoEncontrado(msg)

def _normalize_uf(uf: Optional[str]) -> Optional[str]:
    if not uf:
        return None
    uf = str(uf).strip().upper()
    valid = {u for u, _ in UF_CHOICES}
    return uf if uf in valid else None

# =====================
# CRUD
# =====================
@transaction.atomic
def criar_condominio(data: Dict[str, Any]) -> Condominio:
    if "cnpj" not in data:
        raise ValidationError("Campo 'cnpj' é obrigatório.")
    data = {**data}
    data["cnpj"] = _only_digits(data["cnpj"])
    if Condominio.objects.filter(cnpj=data["cnpj"]).exists():
        raise CondominioJaExiste("Já existe um condomínio com este CNPJ.")
    # normaliza UF
    if "estado" in data:
        data["estado"] = _normalize_uf(data.get("estado")) or ""
    cond = Condominio.objects.create(**data)
    return cond

@transaction.atomic
def atualizar_condominio(condominio_id: int, data: Dict[str, Any]) -> Condominio:
    cond = Condominio.objects.filter(id=condominio_id).first()
    _assert_condominio(cond)
    data = {**data}
    if "cnpj" in data and data["cnpj"]:
        novo = _only_digits(data["cnpj"])
        if Condominio.objects.exclude(id=condominio_id).filter(cnpj=novo).exists():
            raise CondominioJaExiste("Já existe um condomínio com este CNPJ.")
        data["cnpj"] = novo
    if "estado" in data:
        data["estado"] = _normalize_uf(data.get("estado")) or ""
    for k, v in data.items():
        setattr(cond, k, v)
    cond.full_clean()
    cond.save()
    return cond

@transaction.atomic
def inativar_condominio(condominio_id: int) -> Condominio:
    cond = Condominio.objects.filter(id=condominio_id).first()
    _assert_condominio(cond)
    cond.ativo = False
    cond.save(update_fields=["ativo"])
    return cond

@transaction.atomic
def reativar_condominio(condominio_id: int) -> Condominio:
    cond = Condominio.objects.filter(id=condominio_id).first()
    _assert_condominio(cond)
    cond.ativo = True
    cond.save(update_fields=["ativo"])
    return cond

def obter_condominio_por_id(condominio_id: int) -> Condominio:
    cond = Condominio.objects.filter(id=condominio_id).first()
    _assert_condominio(cond)
    return cond

def obter_condominio_por_cnpj(cnpj: str) -> Condominio:
    doc = _only_digits(cnpj)
    cond = Condominio.objects.filter(cnpj=doc).first()
    _assert_condominio(cond)
    return cond

# =====================
# Busca e paginação
# =====================
def buscar_condominios(q: str = "", uf: Optional[str] = None, cidade: str = "", ativos: Optional[bool] = None):
    qs = Condominio.objects.all()
    if q:
        q_digits = _only_digits(q)
        qs = qs.filter(
            Q(nome__icontains=q) |
            Q(email__icontains=q) |
            Q(cnpj__icontains=q_digits)
        )
    if cidade:
        qs = qs.filter(municipio__icontains=cidade.strip())
    if uf:
        norm = _normalize_uf(uf)
        if norm:
            qs = qs.filter(estado=norm)
    if ativos is not None:
        qs = qs.filter(ativo=ativos)
    return qs.order_by("nome", "id")

def paginar_queryset(qs, page: int = 1, per_page: int = 20):
    paginator = Paginator(qs, per_page)
    try:
        return paginator.page(page)
    except PageNotAnInteger:
        return paginator.page(1)
    except EmptyPage:
        return paginator.page(paginator.num_pages if page > 1 else 1)

# =====================
# Atualizações parciais
# =====================
@transaction.atomic
def atualizar_endereco(
    condominio_id: int,
    *,
    cep: Optional[str] = None,
    numero: Optional[str] = None,
    logradouro: Optional[str] = None,
    bairro: Optional[str] = None,
    complemento: Optional[str] = None,
    municipio: Optional[str] = None,
    estado: Optional[str] = None,
) -> Condominio:
    cond = obter_condominio_por_id(condominio_id)
    if cep is not None: cond.cep = cep
    if numero is not None: cond.numero = numero
    if logradouro is not None: cond.logradouro = logradouro
    if bairro is not None: cond.bairro = bairro
    if complemento is not None: cond.complemento = complemento
    if municipio is not None: cond.municipio = municipio
    if estado is not None: cond.estado = _normalize_uf(estado) or ""
    cond.full_clean()
    cond.save()
    return cond

# =====================
# Importação / Exportação Excel
# =====================
def importar_condominios_de_excel(file_or_path, *, sheet_name: str | None = None, strategy: str = "upsert") -> dict:
    """
    Importa condomínios de planilha Excel.
    Colunas reconhecidas (case-insensitive; acentos ignorados):
      - cnpj, nome, nome_completo, email, cep, numero, logradouro, bairro,
        complemento, municipio, cidade, estado, uf, ativo
    Retorna relatório com contagens e erros.
    """
    import unicodedata, re
    from django.core.exceptions import ValidationError

    def norm(s: str) -> str:
        s = str(s or "").strip().lower()
        s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
        s = re.sub(r"[^a-z0-9_/ ]+", " ", s)
        s = re.sub(r"\s+", " ", s)
        return s

    header_map = {
        "cnpj": "cnpj",
        "nome": "nome",
        "nome completo": "nome",
        "nome_completo": "nome",
        "email": "email",
        "cep": "cep",
        "numero": "numero",
        "logradouro": "logradouro",
        "bairro": "bairro",
        "complemento": "complemento",
        "municipio": "municipio",
        "cidade": "municipio",
        "estado": "estado",
        "uf": "estado",
        "ativo": "ativo",
    }

    wb = load_workbook(file_or_path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else (wb.active if sheet_name is None else wb.active)

    headers = [norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_to_field = {}
    for idx, h in enumerate(headers):
        if h in header_map:
            col_to_field[idx] = header_map[h]

    required = {"cnpj", "nome"}
    if not required.issubset(set(col_to_field.values())):
        faltando = required - set(col_to_field.values())
        raise ValidationError(f"Planilha incompleta. Faltam colunas: {', '.join(sorted(faltando))}")

    rel = {"total_linhas": 0, "sucesso": 0, "criados": 0, "atualizados": 0, "erros": []}

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        rel["total_linhas"] += 1
        try:
            data = {}
            for idx, cell in enumerate(row):
                field = col_to_field.get(idx)
                if not field:
                    continue
                val = cell.value
                if field == "cnpj":
                    val = _only_digits(val)
                elif field == "estado":
                    val = _normalize_uf(val) or ""
                elif field == "ativo":
                    s = str(val).strip().lower() if val is not None else ""
                    val = s in ("1","true","t","sim","s","yes","y")
                data[field] = val

            if not data.get("cnpj"):
                raise ValidationError("CNPJ vazio.")
            if strategy == "create":
                cond = criar_condominio(data)
                rel["criados"] += 1
            else:
                # upsert
                doc = data["cnpj"]
                existente = Condominio.objects.filter(cnpj=doc).first()
                if existente:
                    cond = atualizar_condominio(existente.id, data)
                    rel["atualizados"] += 1
                else:
                    cond = criar_condominio(data)
                    rel["criados"] += 1

            rel["sucesso"] += 1
        except Exception as e:
            rel["erros"].append({"linha": i, "erro": str(e)})

    return rel

def exportar_condominios_para_excel(queryset=None) -> tuple[str, bytes]:
    """
    Exporta condomínios para Excel.
    Aba única "Condominios" com colunas:
      cnpj, nome, email, cep, numero, logradouro, bairro, complemento, municipio, estado, ativo, id, created_at, updated_at
    """
    qs = queryset if queryset is not None else Condominio.objects.all().order_by("nome")

    wb = Workbook()
    ws = wb.active
    ws.title = "Condominios"

    headers = [
        "cnpj","nome","email","cep","numero","logradouro","bairro","complemento",
        "municipio","estado","ativo","id","created_at","updated_at"
    ]
    ws.append(headers)

    for c in qs:
        ws.append([
            c.cnpj, c.nome, c.email or "", c.cep or "", c.numero or "", c.logradouro or "",
            c.bairro or "", c.complemento or "", c.municipio or "", c.estado or "",
            bool(c.ativo), c.id,
            c.created_at.strftime("%Y-%m-%d %H:%M:%S") if c.created_at else "",
            c.updated_at.strftime("%Y-%m-%d %H:%M:%S") if c.updated_at else "",
        ])

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"condominios_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return filename, bio.read()


def criar_exemplos_fake(qtd: int = 50) -> int:
    """
    Gera 'qtd' condomínios de exemplo.
    Retorna o número de registros criados (duplicados por CNPJ são ignorados).
    """
    try:
        from faker import Faker
    except ImportError as e:
        raise ImportError("Instale a dependência Faker: pip install Faker") from e

    fake = Faker("pt_BR")
    created = 0
    for _ in range(qtd):
        nome = f"{fake.company()} Condomínio"
        data = {
            "cnpj": "".join(ch for ch in fake.cnpj() if ch.isdigit()),
            "nome": nome,
            "email": fake.company_email(),
            "cep": fake.postcode(),
            "numero": str(fake.random_int(1, 9999)),
            "logradouro": fake.street_name(),
            "bairro": "Centro",
            "complemento": 'SN',
            "municipio": fake.city(),
            "estado": fake.random_element([u for u, _ in UF_CHOICES]),
            "ativo": True,
        }
        try:
            criar_condominio(data)
            created += 1
        except CondominioJaExiste:
            # ignora CNPJ repetido
            continue
    return created
