# -*- coding: utf-8 -*-
from __future__ import annotations
from typing import Optional, Dict, Any
from datetime import datetime, date
from io import BytesIO
from decimal import Decimal

from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.db import transaction
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from .models import Funcionario

class FuncionarioJaExiste(ValidationError): ...
class FuncionarioNaoEncontrado(ObjectDoesNotExist): ...

# ========= Helpers =========
def _digits(s: Any) -> str:
    return "".join(ch for ch in str(s or "") if ch.isdigit())

def paginar_queryset(qs, page: int = 1, per_page: int = 20):
    paginator = Paginator(qs, per_page)
    try:
        return paginator.page(page)
    except PageNotAnInteger:
        return paginator.page(1)
    except EmptyPage:
        return paginator.page(paginator.num_pages if page > 1 else 1)

# ========= CRUD =========
@transaction.atomic
def criar_funcionario(data: Dict[str, Any]) -> Funcionario:
    if not data.get("cpf_cnpj"):
        raise ValidationError("Campo 'cpf_cnpj' é obrigatório.")
    doc = _digits(data["cpf_cnpj"])
    if Funcionario.objects.filter(cpf_cnpj=doc).exists():
        raise FuncionarioJaExiste("Já existe funcionário com este CPF/CNPJ.")

    obj = Funcionario(
        cpf_cnpj=doc,
        nome=data.get("nome") or "",
        email=data.get("email") or "",
        telefone=data.get("telefone") or "",
        rg=data.get("rg") or "",
        registro_cref=data.get("registro_cref") or "",
        tam_uniforme=data.get("tam_uniforme") or "",
        data_nascimento=data.get("data_nascimento") or None,
        data_admissao=data.get("data_admissao") or None,
        cargo=data.get("cargo") or Funcionario.Cargo.OUTRO,
        regime_trabalhista=data.get("regime_trabalhista") or Funcionario.RegimeTrabalhista.OUTRO,
        ativo=bool(data.get("ativo", True)),
    )
    obj.full_clean()
    obj.save()
    return obj

@transaction.atomic
def atualizar_funcionario(funcionario_id: int, data: Dict[str, Any]) -> Funcionario:
    fun = Funcionario.objects.filter(id=funcionario_id).first()
    if not fun:
        raise FuncionarioNaoEncontrado("Funcionário não encontrado.")

    if "cpf_cnpj" in data and data["cpf_cnpj"]:
        novo = _digits(data["cpf_cnpj"])
        if Funcionario.objects.exclude(id=funcionario_id).filter(cpf_cnpj=novo).exists():
            raise FuncionarioJaExiste("Já existe funcionário com este CPF/CNPJ.")
        fun.cpf_cnpj = novo

    for k in (
        "nome", "email", "telefone", "ativo", "cargo", "regime_trabalhista",
        "data_nascimento", "data_admissao", "rg", "registro_cref", "tam_uniforme"
    ):
        if k in data:
            setattr(fun, k, data[k])

    fun.full_clean()
    fun.save()
    return fun

# ========= Inativar/Reativar =========
@transaction.atomic
def inativar_funcionario(funcionario_id: int) -> Funcionario:
    fun = Funcionario.objects.filter(id=funcionario_id).first()
    if not fun:
        raise FuncionarioNaoEncontrado("Funcionário não encontrado.")
    fun.ativo = False
    fun.save(update_fields=["ativo"])
    return fun

@transaction.atomic
def reativar_funcionario(funcionario_id: int) -> Funcionario:
    fun = Funcionario.objects.filter(id=funcionario_id).first()
    if not fun:
        raise FuncionarioNaoEncontrado("Funcionário não encontrado.")
    fun.ativo = True
    fun.save(update_fields=["ativo"])
    return fun

def obter_funcionario_por_id(funcionario_id: int) -> Funcionario:
    fun = Funcionario.objects.filter(id=funcionario_id).first()
    if not fun:
        raise FuncionarioNaoEncontrado("Funcionário não encontrado.")
    return fun

def obter_funcionario_por_doc(cpf_cnpj: str) -> Funcionario:
    doc = _digits(cpf_cnpj)
    fun = Funcionario.objects.filter(cpf_cnpj=doc).first()
    if not fun:
        raise FuncionarioNaoEncontrado("Funcionário não encontrado.")
    return fun

# ========= Busca =========
def buscar_funcionarios(q: str = "", ativo: bool | None = None, regime: str | None = None, cargo: str | None = None):
    qs = Funcionario.objects.all()
    if q:
        qs = qs.filter(
            Q(nome__icontains=q) |
            Q(email__icontains=q) |
            Q(cpf_cnpj__icontains=q)
        )
    if ativo is not None:
        qs = qs.filter(ativo=ativo)
    if regime:
        qs = qs.filter(regime_trabalhista=regime)
    if cargo:
        qs = qs.filter(cargo=cargo)
    return qs.order_by("nome", "id")

# ========= Importação / Exportação Excel =========
def importar_funcionarios_de_excel(file_or_path, *, sheet_name: str | None = None, strategy: str = "upsert") -> dict:
    import unicodedata, re
    def norm(s: str) -> str:
        s = str(s or "").strip().lower()
        s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
        s = re.sub(r"[^a-z0-9_/ ]+", " ", s)
        s = re.sub(r"\s+", " ", s)
        return s

    header_map = {
        "cpf": "cpf_cnpj",
        "cpf_cnpj": "cpf_cnpj",
        "cnpj": "cpf_cnpj",
        "nome": "nome",
        "email": "email",
        "telefone": "telefone",
        "ativo": "ativo",
        "cargo": "cargo",
        "data_nascimento": "data_nascimento",
    }

    wb = load_workbook(file_or_path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else (wb.active if sheet_name is None else wb.active)

    headers = [norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_to_field = {idx: header_map[h] for idx, h in enumerate(headers) if h in header_map}

    required = {"cpf_cnpj", "nome"}
    if not required.issubset(set(col_to_field.values())):
        faltando = required - set(col_to_field.values())
        raise ValidationError(f"Planilha incompleta. Faltam colunas: {', '.join(sorted(faltando))}")

    rel = {"total_linhas": 0, "sucesso": 0, "criados": 0, "atualizados": 0, "erros": []}

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        rel["total_linhas"] += 1
        try:
            data = {}
            for idx, cell in enumerate(row):
                field = col_to_field.get(idx)
                if not field:
                    continue
                val = cell.value
                if field == "cpf_cnpj":
                    val = _digits(val)
                elif field == "ativo":
                    s = str(val).strip().lower() if val is not None else ""
                    val = s in ("1","true","t","sim","s","yes","y")
                elif field == "data_nascimento" and val:
                    try:
                        val = val if isinstance(val, (date, datetime)) else datetime.strptime(str(val), "%Y-%m-%d").date()
                    except:
                        val = None
                data[field] = val

            if not data.get("cpf_cnpj"):
                raise ValidationError("CPF/CNPJ vazio.")

            doc = data["cpf_cnpj"]
            existente = Funcionario.objects.filter(cpf_cnpj=doc).first()
            if strategy == "create" or not existente:
                criar_funcionario(data)
                rel["criados"] += 1
            else:
                atualizar_funcionario(existente.id, data)
                rel["atualizados"] += 1

            rel["sucesso"] += 1
        except Exception as e:
            rel["erros"].append({"linha": i, "erro": str(e)})

    return rel

def exportar_funcionarios_para_excel(queryset=None) -> tuple[str, bytes]:
    qs = queryset if queryset is not None else Funcionario.objects.all().order_by("nome", "id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Funcionarios"

    headers = [
        "cpf_cnpj","nome","email","telefone","rg","registro_cref",
        "tam_uniforme","data_nascimento","data_admissao","cargo",
        "regime_trabalhista","ativo","id","created_at","updated_at"
    ]
    ws.append(headers)

    for f in qs:
        ws.append([
            f.cpf_cnpj,
            f.nome,
            f.email or "",
            f.telefone or "",
            f.rg or "",
            f.registro_cref or "",
            f.tam_uniforme or "",
            f.data_nascimento.strftime("%d/%m/%Y") if f.data_nascimento else "",
            f.data_admissao.strftime("%d/%m/%Y") if f.data_admissao else "",
            f.get_cargo_display(),
            f.get_regime_trabalhista_display(),
            bool(f.ativo),
            f.id,
            f.created_at.strftime("%Y-%m-%d %H:%M:%S") if f.created_at else "",
            f.updated_at.strftime("%Y-%m-%d %H:%M:%S") if f.updated_at else "",
        ])

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"funcionarios_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return filename, bio.read()
# ========= Dados Fake =========
def criar_exemplos_fake(qtd: int = 40) -> int:
    try:
        from faker import Faker
    except ImportError:
        raise ImportError("Instale Faker: pip install Faker")
    fake = Faker("pt_BR")
    created = 0
    for _ in range(qtd):
        doc = "".join(ch for ch in fake.cpf() if ch.isdigit())
        data = {
            "cpf_cnpj": doc,
            "nome": fake.name(),
            "email": fake.free_email(),
            "telefone": fake.phone_number(),
            "cargo": Funcionario.Cargo.OUTRO,
            "data_nascimento": fake.date_of_birth(minimum_age=18, maximum_age=60),
            "ativo": True,
        }
        try:
            criar_funcionario(data)
            created += 1
        except FuncionarioJaExiste:
            continue
    return created
