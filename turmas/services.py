from __future__ import annotations

from datetime import date, time
from decimal import Decimal
from typing import Dict, Optional, Tuple, List, Any

from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.db import transaction
from django.db.models import Q, QuerySet

from .models import Turma

# Mapa de dia da semana
_DIA_FIELD = {1: "seg", 2: "ter", 3: "qua", 4: "qui", 5: "sex", 6: "sab", 7: "dom"}

def _dias_ativos_from_obj(t: Turma) -> List[str]:
    out = []
    if t.seg: out.append("Seg")
    if t.ter: out.append("Ter")
    if t.qua: out.append("Qua")
    if t.qui: out.append("Qui")
    if t.sex: out.append("Sex")
    if t.sab: out.append("S√°b")
    if t.dom: out.append("Dom")
    return out

def _flags_from_data(data: Dict[str, Any], fallback: Optional[Turma] = None) -> Dict[str, bool]:
    def getb(k: str) -> bool:
        if k in data:
            v = data[k]
            if isinstance(v, bool):
                return v
            return str(v).lower() in ("1", "true", "on")
        return bool(getattr(fallback, k)) if fallback else False

    return {k: getb(k) for k in ("seg", "ter", "qua", "qui", "sex", "sab", "dom")}


# Valida√ß√£o de conflitos de hor√°rio/professor
def _checar_conflito_professor(
    professor_id: int,
    flags: Dict[str, bool],
    hora_inicio: time,
    duracao_minutos: int,
    inicio_vigencia: date,
    fim_vigencia: Optional[date],
    turma_id_excluir: Optional[int] = None,
):
    if sum(flags.values()) == 0:
        raise ValidationError("Selecione ao menos um dia da semana.")

    qs = Turma.objects.filter(professor_id=professor_id)
    if turma_id_excluir:
        qs = qs.exclude(id=turma_id_excluir)

    dia_q = Q()
    for k, ativo in flags.items():
        if ativo:
            dia_q |= Q(**{k: True})
    qs = qs.filter(dia_q)

    qs = qs.filter(
        Q(fim_vigencia__isnull=True, inicio_vigencia__lte=fim_vigencia or date.max) |
        Q(
            fim_vigencia__isnull=False,
            inicio_vigencia__lte=fim_vigencia or date.max,
            fim_vigencia__gte=inicio_vigencia,
        )
    )

    def _hora_cruza(hini: time, dur: int, hini_b: time, dur_b: int) -> bool:
        a0, a1 = hini.hour * 60 + hini.minute, hini.hour * 60 + hini.minute + dur
        b0, b1 = hini_b.hour * 60 + hini_b.minute, hini_b.hour * 60 + hini_b.minute + dur_b
        return a0 < b1 and b0 < a1

    for t in qs:
        if _hora_cruza(hora_inicio, duracao_minutos, t.hora_inicio, t.duracao_minutos):
            raise ValidationError("Conflito de hor√°rio para o professor.")


# --------------------- CRUD ------------------------

@transaction.atomic
def criar_turma(data: Dict[str, Any]) -> Turma:
    flags = _flags_from_data(data)

    professor = data["professor"]
    professor_id = professor.id if hasattr(professor, "id") else int(professor)

    _checar_conflito_professor(
        professor_id=professor_id,
        flags=flags,
        hora_inicio=data["hora_inicio"],
        duracao_minutos=int(data["duracao_minutos"]),
        inicio_vigencia=data["inicio_vigencia"],
        fim_vigencia=data.get("fim_vigencia"),
    )

    if data.get("fim_vigencia") and data["fim_vigencia"] < data["inicio_vigencia"]:
        raise ValidationError("A data de fim n√£o pode ser antes do in√≠cio da vig√™ncia.")

    return Turma.objects.create(**data)


@transaction.atomic
def atualizar_turma(turma_id: int, data: Dict[str, Any]) -> Turma:
    t = Turma.objects.filter(id=turma_id).first()
    if not t:
        raise ObjectDoesNotExist("Turma n√£o encontrada.")

    flags = _flags_from_data(data, fallback=t)

    professor_data = data.get("professor", t.professor_id)
    professor_id = professor_data.id if hasattr(professor_data, "id") else int(professor_data)

    _checar_conflito_professor(
        professor_id=professor_id,
        flags=flags,
        hora_inicio=data.get("hora_inicio", t.hora_inicio),
        duracao_minutos=int(data.get("duracao_minutos", t.duracao_minutos)),
        inicio_vigencia=data.get("inicio_vigencia", t.inicio_vigencia),
        fim_vigencia=data.get("fim_vigencia", t.fim_vigencia),
        turma_id_excluir=t.id,
    )

    for k, v in data.items():
        setattr(t, k, v)

    if t.fim_vigencia and t.fim_vigencia < t.inicio_vigencia:
        raise ValidationError("A data de fim da vig√™ncia n√£o pode ser anterior ao in√≠cio.")

    t.full_clean()
    t.save()
    return t


from django.db import transaction
from django.db.models import Q, QuerySet
from django.core.exceptions import ValidationError, ObjectDoesNotExist
from datetime import datetime
from decimal import Decimal
from typing import Optional, Tuple, List, Any, Dict
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Turma
from .models import Matricula
from clientes.models import Cliente

# ------------------------------------------------------------
# Matr√≠culas
# ------------------------------------------------------------
from django.core.exceptions import ValidationError

def matricular_cliente(
    turma_id: int,
    cliente_id: int,
    data_inicio: date,
    participante_nome: str | None = None,
    participante_data_nascimento: date | None = None,
    participante_sexo: str | None = None,
    proprio_cliente: bool = True,
):
    """
    Cria uma matr√≠cula para o cliente ou dependente.
    Usa a l√≥gica de capacidade e ocupa√ß√£o do model Turma.
    """

    turma = Turma.objects.get(id=turma_id)
    cliente = Cliente.objects.get(id=cliente_id)

    print(
        f"Tentando matricular: turma={turma}, cliente={cliente}, "
        f"proprio_cliente={proprio_cliente}, participante={participante_nome}"
    )

    # üö´ Checa se a turma j√° est√° lotada (usa a propriedade do model)
    if turma.lotada:
        raise ValidationError(
            f"A turma '{turma}' j√° atingiu sua capacidade m√°xima de {turma.capacidade} alunos."
        )

    # üîç 1. Caso seja o pr√≥prio cliente
    if proprio_cliente:
        participante_nome = ""
        participante_data_nascimento = None
        participante_sexo = ""

        # Evita duplicidade do titular
        if Matricula.objects.filter(
            turma=turma, cliente=cliente, participante_nome="", ativa=True
        ).exists():
            raise ValidationError("Este cliente j√° est√° matriculado nesta turma.")

    # üîç 2. Caso seja dependente
    else:
        if not participante_nome:
            raise ValidationError("Informe o nome do dependente.")

        # Evita duplicidade do mesmo dependente
        if Matricula.objects.filter(
            turma=turma,
            cliente=cliente,
            participante_nome__iexact=participante_nome.strip(),
            ativa=True,
        ).exists():
            raise ValidationError(
                f"O dependente '{participante_nome}' j√° est√° matriculado nesta turma."
            )

    # ‚úÖ Cria a matr√≠cula
    matricula = Matricula.objects.create(
        turma=turma,
        cliente=cliente,
        data_inicio=data_inicio,
        participante_nome=(participante_nome or "").strip(),
        participante_data_nascimento=participante_data_nascimento,
        participante_sexo=(participante_sexo or "").strip(),
        ativa=True,
    )

    return matricula




@transaction.atomic
def desmatricular(matricula_id: int):
    from .models import Matricula

    m = Matricula.objects.filter(id=matricula_id).first()
    if not m:
        raise ObjectDoesNotExist("Matr√≠cula n√£o encontrada.")
    m.ativa = False
    m.save(update_fields=["ativa"])
    return m


# ------------------------------------------------------------
# Buscar Turmas
# ------------------------------------------------------------
def buscar_turmas(
    *,
    q: str = "",
    condominio_id: Optional[int] = None,
    modalidade_id: Optional[int] = None,
    professor_id: Optional[int] = None,
    dia_semana: Optional[int] = None,
    ativos: Optional[bool] = None,
) -> QuerySet[Turma]:
    qs = Turma.objects.select_related("modalidade__condominio", "professor").all()

    if q:
        qs = qs.filter(
            Q(nome_exibicao__icontains=q) |
            Q(modalidade__nome__icontains=q) |
            Q(modalidade__condominio__nome__icontains=q) |
            Q(professor__nome__icontains=q)
        )

    if condominio_id:
        qs = qs.filter(modalidade__condominio_id=condominio_id)

    if modalidade_id:
        qs = qs.filter(modalidade_id=modalidade_id)

    if professor_id:
        qs = qs.filter(professor_id=professor_id)

    if dia_semana not in (None, ""):
        try:
            field = _DIA_FIELD[int(dia_semana)]
            qs = qs.filter(**{field: True})
        except Exception:
            pass

    if ativos is not None:
        qs = qs.filter(ativo=bool(ativos))

    return qs.order_by("modalidade__condominio__nome", "modalidade__nome", "hora_inicio", "id")


# ------------------------------------------------------------
# Exporta√ß√£o de Turmas para Excel (com campos novos)
# ------------------------------------------------------------
def exportar_turmas_excel(qs: Optional[QuerySet[Turma]] = None) -> Tuple[str, bytes]:
    qs = qs or Turma.objects.select_related("modalidade__condominio", "professor").all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Turmas"

    headers = [
        "ID", "Modalidade", "Condom√≠nio", "Professor", "Dias",
        "Hora In√≠cio", "Dura√ß√£o (min)", "Capacidade",
        "Valor", "Valor + DSR", "VT/VA", "Bonifica√ß√£o",
        "Observa√ß√µes", "Ativa", "In√≠cio Vig√™ncia", "Fim Vig√™ncia", "Nome Exibi√ß√£o"
    ]
    ws.append(headers)

    for t in qs:
        dias = ", ".join(_dias_ativos_from_obj(t)) or "‚Äî"
        ws.append([
            t.id,
            getattr(t.modalidade, "nome", ""),
            getattr(t.modalidade.condominio, "nome", "") if t.modalidade else "",
            getattr(t.professor, "nome", ""),
            dias,
            t.hora_inicio.strftime("%H:%M"),
            t.duracao_minutos,
            t.capacidade,
            f"{t.valor:.2f}",
            f"{getattr(t, 'valor_dsr', Decimal('0.00')):.2f}",
            f"{getattr(t, 'vt_va', Decimal('0.00')):.2f}",
            f"{getattr(t, 'bonificacao', Decimal('0.00')):.2f}",
            t.obs or "",
            "SIM" if t.ativo else "N√ÉO",
            t.inicio_vigencia.strftime("%d/%m/%Y"),
            t.fim_vigencia.strftime("%d/%m/%Y") if t.fim_vigencia else "",
            t.nome_exibicao or "",
        ])

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"turmas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return filename, bio.read()
