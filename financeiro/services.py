from __future__ import annotations
from typing import Optional, Dict, Any, Iterable, Tuple
from datetime import date, timedelta
from decimal import Decimal
from calendar import monthrange

from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.db import transaction
from django.db.models import Q, Sum, F

from .models import Lancamento, Baixa, CategoriaFinanceira  # financeiro
# ===== Helpers =====
def paginar_queryset(qs, page: int = 1, per_page: int = 20):
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    p = Paginator(qs, per_page)
    try:
        return p.page(page)
    except PageNotAnInteger:
        return p.page(1)
    except EmptyPage:
        return p.page(p.num_pages if page > 1 else 1)

def _atualizar_status(l: Lancamento):
    if l.status == "CANCELADO":
        return
    saldo = l.saldo
    if saldo <= Decimal("0.00"):
        l.status = "LIQUIDADO"
    elif l.total_baixado > Decimal("0.00"):
        l.status = "PARCIAL"
    else:
        l.status = "ABERTO"
    l.save(update_fields=["status", "updated_at"])

# ===== CRUD Lançamento =====
@transaction.atomic
def criar_lancamento(data: Dict[str, Any]) -> Lancamento:
    l = Lancamento.objects.create(**data)
    _atualizar_status(l)
    return l

@transaction.atomic
def atualizar_lancamento(lancamento_id: int, data: Dict[str, Any]) -> Lancamento:
    l = Lancamento.objects.filter(id=lancamento_id).first()
    if not l:
        raise ObjectDoesNotExist("Lançamento não encontrado.")
    # não permitir alterar valor para abaixo do já baixado
    if "valor" in data:
        novo_valor = Decimal(str(data["valor"]))
        if novo_valor < l.total_baixado:
            raise ValidationError("Valor não pode ser inferior ao total já baixado.")
    for k, v in data.items():
        setattr(l, k, v)
    l.full_clean()
    l.save()
    _atualizar_status(l)
    return l

@transaction.atomic
def cancelar_lancamento(lancamento_id: int) -> Lancamento:
    l = Lancamento.objects.filter(id=lancamento_id).first()
    if not l:
        raise ObjectDoesNotExist("Lançamento não encontrado.")
    if l.baixas.exists():
        raise ValidationError("Não é possível cancelar: já possui baixas.")
    l.status = "CANCELADO"
    l.ativo = False
    l.save(update_fields=["status", "ativo", "updated_at"])
    return l

# ===== Baixas =====
@transaction.atomic
def registrar_baixa(*, lancamento_id: int, valor: Decimal, data: date, forma: str, observacao: str = "") -> Baixa:
    l = Lancamento.objects.select_for_update().filter(id=lancamento_id).first()
    if not l:
        raise ObjectDoesNotExist("Lançamento não encontrado.")
    if l.status == "CANCELADO":
        raise ValidationError("Lançamento cancelado.")
    if valor <= Decimal("0"):
        raise ValidationError("Valor da baixa deve ser > 0.")
    if valor > l.saldo:
        raise ValidationError("Valor maior que o saldo do lançamento.")

    b = Baixa.objects.create(lancamento=l, valor=valor, data=data, forma=forma, observacao=observacao)
    _atualizar_status(l)
    return b

@transaction.atomic
def estornar_baixa(baixa_id: int) -> Lancamento:
    b = Baixa.objects.select_related("lancamento").filter(id=baixa_id).first()
    if not b:
        raise ObjectDoesNotExist("Baixa não encontrada.")
    l = b.lancamento
    b.delete()
    _atualizar_status(l)
    return l

# ===== Busca/Relatórios simples =====
def buscar_lancamentos(
    *,
    q: str = "",
    tipo: Optional[str] = None,
    status: Optional[str] = None,
    venc_de: Optional[date] = None,
    venc_ate: Optional[date] = None,
    cliente_id: Optional[int] = None,
    funcionario_id: Optional[int] = None,
    condominio_id: Optional[int] = None,
    turma_id: Optional[int] = None,
    categoria_id: Optional[int] = None,
    ativos: Optional[bool] = None,
):
    qs = (Lancamento.objects
          .select_related("cliente", "funcionario", "condominio", "turma", "categoria")
          .all())
    if q:
        qs = qs.filter(
            Q(descricao__icontains=q) |
            Q(observacao__icontains=q) |
            Q(contraparte_nome__icontains=q)
        )
    if tipo: qs = qs.filter(tipo=tipo)
    if status: qs = qs.filter(status=status)
    if venc_de: qs = qs.filter(vencimento__gte=venc_de)
    if venc_ate: qs = qs.filter(vencimento__lte=venc_ate)
    if cliente_id: qs = qs.filter(cliente_id=cliente_id)
    if funcionario_id: qs = qs.filter(funcionario_id=funcionario_id)
    if condominio_id: qs = qs.filter(condominio_id=condominio_id)
    if turma_id: qs = qs.filter(turma_id=turma_id)
    if categoria_id: qs = qs.filter(categoria_id=categoria_id)
    if ativos is not None: qs = qs.filter(ativo=ativos)

    qs = qs.annotate(total_baixado_agg=Sum("baixas__valor")).order_by("-vencimento", "-id")
    return qs

# ===== Recorrência mensal (manual) =====
@transaction.atomic
def gerar_recorrencia_mensal(
    *,
    tipo: str,
    descricao: str,
    valor: Decimal,
    dia_venc: int,
    quantidade: int,
    primeiro_mes: date,
    relacionamentos: Dict[str, int] = None,
    categoria_id: Optional[int] = None,
    observacao: str = "",
) -> int:
    """
    Gera N lançamentos mensais com mesmo valor/dia de vencimento.
    `relacionamentos` pode conter ids: cliente_id, funcionario_id, condominio_id, turma_id.
    Retorna quantidade criada.
    """
    if tipo not in ("RECEBER", "PAGAR"):
        raise ValidationError("Tipo inválido.")
    if quantidade < 1:
        raise ValidationError("Quantidade deve ser >= 1.")
    relacionamentos = relacionamentos or {}
    criados = 0

    ano = primeiro_mes.year
    mes = primeiro_mes.month

    from calendar import monthrange
    for _ in range(quantidade):
        last_day = monthrange(ano, mes)[1]
        dia = min(dia_venc, last_day)
        venc = date(ano, mes, dia)

        Lancamento.objects.create(
            tipo=tipo,
            descricao=descricao,
            valor=valor,
            vencimento=venc,
            status="ABERTO",
            cliente_id=relacionamentos.get("cliente_id"),
            funcionario_id=relacionamentos.get("funcionario_id"),
            condominio_id=relacionamentos.get("condominio_id"),
            turma_id=relacionamentos.get("turma_id"),
            categoria_id=categoria_id,
            observacao=observacao,
        )
        criados += 1

        # próximo mês
        if mes == 12:
            mes = 1; ano += 1
        else:
            mes += 1

    return criados

# ===== Exportação Excel =====
def exportar_lancamentos_excel(queryset=None) -> tuple[str, bytes]:
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    qs = queryset if queryset is not None else Lancamento.objects.all().select_related(
        "cliente","funcionario","condominio","turma","categoria"
    ).order_by("-vencimento","-id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Financeiro"

    headers = ["Tipo","Descrição","Vencimento","Valor","Baixado","Saldo","Status",
               "Cliente","Funcionário","Condomínio","Turma","Categoria","Obs","ID"]
    ws.append(headers)

    for l in qs:
        total_baixado = l.total_baixado
        ws.append([
            l.get_tipo_display(),
            l.descricao,
            l.vencimento.strftime("%Y-%m-%d"),
            float(l.valor),
            float(total_baixado),
            float(l.saldo),
            l.get_status_display(),
            (l.cliente.nome_razao if l.cliente_id else l.contraparte_nome),
            (l.funcionario.nome if l.funcionario_id else ""),
            (l.condominio.nome if l.condominio_id else ""),
            (l.turma.nome_exibicao or str(l.turma) if l.turma_id else ""),
            (l.categoria.nome if l.categoria_id else ""),
            (l.observacao or "")[:250],
            l.id,
        ])

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    bio = BytesIO()
    wb.save(bio); bio.seek(0)

    from datetime import datetime
    filename = f"financeiro_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return filename, bio.read()

# === Cobranças de Mensalidades (AGORA agregadas por CLIENTE) ===
from datetime import date as _date
from turmas.models import Turma, Matricula

def _primeiro_ultimo_dia(ano: int, mes: int) -> tuple[_date, _date]:
    last = monthrange(ano, mes)[1]
    return _date(ano, mes, 1), _date(ano, mes, last)

def _clamp_vencimento(ano: int, mes: int, dia_venc: int) -> _date:
    last = monthrange(ano, mes)[1]
    return _date(ano, mes, min(dia_venc, last))

def _matriculas_ativas_no_mes_global(ano: int, mes: int):
    """Todas matrículas ativas que cruzam a competência (qualquer dia no mês)."""
    inicio_mes, fim_mes = _primeiro_ultimo_dia(ano, mes)
    return (Matricula.objects
            .select_related("cliente", "turma", "turma__modalidade", "turma__condominio")
            .filter(ativa=True)
            .filter(Q(data_fim__isnull=True) | Q(data_fim__gte=inicio_mes))
            .filter(data_inicio__lte=fim_mes))

def _matriculas_ativas_no_mes_da_turma(turma: Turma, ano: int, mes: int):
    """Matrículas ativas no mês, apenas da turma informada."""
    inicio_mes, fim_mes = _primeiro_ultimo_dia(ano, mes)
    return (Matricula.objects
            .select_related("cliente", "turma", "turma__modalidade", "turma__condominio")
            .filter(turma=turma, ativa=True)
            .filter(Q(data_fim__isnull=True) | Q(data_fim__gte=inicio_mes))
            .filter(data_inicio__lte=fim_mes))

def _get_or_create_categoria(nome: str = "Mensalidades") -> CategoriaFinanceira:
    cat, _ = CategoriaFinanceira.objects.get_or_create(nome=nome)
    return cat

def _ja_existe_cobranca_cliente_mes(cliente_id: int, ano: int, mes: int) -> bool:
    """Evita duplicidade por (cliente, mês). Ignora CANCELADO."""
    return Lancamento.objects.filter(
        tipo="RECEBER",
        cliente_id=cliente_id,
        vencimento__year=ano,
        vencimento__month=mes
    ).exclude(status="CANCELADO").exists()

def _desconto_percent_por_modalidades(qtd_modalidades: int) -> Decimal:
    if qtd_modalidades >= 4:
        return Decimal("0.10")
    if qtd_modalidades == 3:
        return Decimal("0.075")
    if qtd_modalidades == 2:
        return Decimal("0.05")
    return Decimal("0.00")

def _agrupar_totais_por_cliente(mats: Iterable[Matricula]) -> Dict[int, Dict[str, Any]]:
    """
    Retorna um dict:
      { cliente_id: {
            "cliente": <obj>,
            "modalidades": { modalidade_id: {"nome": str, "qtd": int, "valor_unit": Decimal} },
            "subtotal": Decimal,
            "qtd_modalidades": int
        }}
    """
    grouped: Dict[int, Dict[str, Any]] = {}
    for m in mats:
        cid = m.cliente_id
        t = m.turma
        mod = t.modalidade
        valor_unit = Decimal(t.valor)  # valor da modalidade (na turma)
        entry = grouped.setdefault(cid, {
            "cliente": m.cliente,
            "modalidades": {},  # por modalidade
            "subtotal": Decimal("0.00"),
            "qtd_modalidades": 0,
        })
        mod_entry = entry["modalidades"].setdefault(mod.id, {"nome": mod.nome, "qtd": 0, "valor_unit": valor_unit})
        mod_entry["qtd"] += 1
    # calcula subtotais e contagem de modalidades
    for cid, entry in grouped.items():
        subtotal = Decimal("0.00")
        modalidade_ids = []
        for mid, info in entry["modalidades"].items():
            modalidade_ids.append(mid)
            subtotal += Decimal(info["qtd"]) * Decimal(info["valor_unit"])
        entry["subtotal"] = subtotal
        entry["qtd_modalidades"] = len(set(modalidade_ids))
    return grouped

@transaction.atomic
def gerar_cobrancas_mensalidade_turma(
    *,
    turma_id: int,
    ano: int,
    mes: int,
    dia_venc: int = 5,
    categoria_nome: str = "Mensalidades",
    descricao_tpl: str = "Mensalidades ({mes:02d}/{ano})",
    observacao_padrao: str = "Gerado automaticamente por turma (agregado por cliente)",
) -> dict:
    """
    Agora cria UMA cobrança por cliente (no mês), considerando TODAS as suas matrículas
    ativas no mês (não apenas desta turma) — para evitar múltiplas cobranças separadas.
    Se já existir cobrança daquele cliente no mês, não cria novamente.
    Retorna: {"criados": X, "existentes": Y, "turma": turma_id, "vencimento": date}
    """
    turma = Turma.objects.select_related("modalidade", "condominio").filter(id=turma_id, ativo=True).first()
    if not turma:
        raise ValueError("Turma não encontrada ou inativa.")

    # se mês fora da vigência da turma, não cria nada (consistente com antes)
    inicio_mes, fim_mes = _primeiro_ultimo_dia(ano, mes)
    if not (turma.inicio_vigencia <= fim_mes and (turma.fim_vigencia is None or turma.fim_vigencia >= inicio_mes)):
        return {"criados": 0, "existentes": 0, "turma": turma_id, "vencimento": None}

    cat = _get_or_create_categoria(categoria_nome)
    venc = _clamp_vencimento(ano, mes, dia_venc)

    # clientes que têm matrícula nesta turma (no mês)
    clientes_ids = set(m.cliente_id for m in _matriculas_ativas_no_mes_da_turma(turma, ano, mes))

    # mas a cobrança é pelo conjunto de TODAS as matrículas do cliente (no mês)
    mats_todas = _matriculas_ativas_no_mes_global(ano, mes).filter(cliente_id__in=clientes_ids)

    por_cliente = _agrupar_totais_por_cliente(mats_todas)

    criados = existentes = 0
    for cid, dados in por_cliente.items():
        if _ja_existe_cobranca_cliente_mes(cid, ano, mes):
            existentes += 1
            continue

        subtotal = dados["subtotal"]
        qtd_modalidades = dados["qtd_modalidades"]
        desc_pct = _desconto_percent_por_modalidades(qtd_modalidades)
        desconto = (subtotal * desc_pct).quantize(Decimal("0.01"))
        total = (subtotal - desconto).quantize(Decimal("0.01"))

        cliente = dados["cliente"]
        descricao = descricao_tpl.format(ano=ano, mes=mes)

        # Observação com breakdown
        parts = []
        for info in dados["modalidades"].values():
            parts.append(f"{info['nome']} x{info['qtd']} @ {Decimal(info['valor_unit']):.2f}")
        breakdown = "; ".join(parts)
        obs = (f"{observacao_padrao}. competência={ano}-{mes:02d}; "
               f"modalidades={qtd_modalidades}; desconto={desc_pct*Decimal('100')}%; "
               f"itens=[{breakdown}]")

        Lancamento.objects.create(
            tipo="RECEBER",
            descricao=f"{descricao} — {cliente.nome_razao}",
            valor=total,
            vencimento=venc,
            status="ABERTO",
            cliente_id=cid,
            turma_id=None,  # agregado por cliente (não por turma)
            categoria_id=cat.id,
            observacao=obs,
            contraparte_nome=cliente.nome_razao,
            contraparte_doc=cliente.cpf_cnpj,
        )
        criados += 1

    return {"criados": criados, "existentes": existentes, "turma": turma_id, "vencimento": venc}

@transaction.atomic
def gerar_cobrancas_mensalidades_global(
    *,
    ano: int,
    mes: int,
    dia_venc: int = 5,
    categoria_nome: str = "Mensalidades",
    descricao_tpl: str = "Mensalidades ({mes:02d}/{ano})",
    observacao_padrao: str = "Gerado automaticamente (global, agregado por cliente)",
) -> dict:
    """
    Cria UMA cobrança por cliente (no mês), considerando TODAS as matrículas ativas.
    Se já existir cobrança daquele cliente no mês, não cria novamente.
    Retorna: {"criados": X, "existentes": Y, "ano": ano, "mes": mes, "dia_venc": dia_venc}
    """
    cat = _get_or_create_categoria(categoria_nome)
    venc = _clamp_vencimento(ano, mes, dia_venc)

    mats = _matriculas_ativas_no_mes_global(ano, mes)
    por_cliente = _agrupar_totais_por_cliente(mats)

    criados = existentes = 0
    for cid, dados in por_cliente.items():
        if _ja_existe_cobranca_cliente_mes(cid, ano, mes):
            existentes += 1
            continue

        subtotal = dados["subtotal"]
        qtd_modalidades = dados["qtd_modalidades"]
        desc_pct = _desconto_percent_por_modalidades(qtd_modalidades)
        desconto = (subtotal * desc_pct).quantize(Decimal("0.01"))
        total = (subtotal - desconto).quantize(Decimal("0.01"))

        cliente = dados["cliente"]
        descricao = descricao_tpl.format(ano=ano, mes=mes)

        parts = []
        for info in dados["modalidades"].values():
            parts.append(f"{info['nome']} x{info['qtd']} @ {Decimal(info['valor_unit']):.2f}")
        breakdown = "; ".join(parts)
        obs = (f"{observacao_padrao}. competência={ano}-{mes:02d}; "
               f"modalidades={qtd_modalidades}; desconto={desc_pct*Decimal('100')}%; "
               f"itens=[{breakdown}]")

        Lancamento.objects.create(
            tipo="RECEBER",
            descricao=f"{descricao} — {cliente.nome_razao}",
            valor=total,
            vencimento=venc,
            status="ABERTO",
            cliente_id=cid,
            turma_id=None,
            categoria_id=cat.id,
            observacao=obs,
            contraparte_nome=cliente.nome_razao,
            contraparte_doc=cliente.cpf_cnpj,
        )
        criados += 1

    return {"criados": criados, "existentes": existentes, "ano": ano, "mes": mes, "dia_venc": dia_venc}
